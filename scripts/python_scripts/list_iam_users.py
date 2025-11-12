import boto3
import pandas as pd
from datetime import datetime, timezone
from tabulate import tabulate
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
import os

iam = boto3.client('iam')

def days_since(date):
    if not date:
        return None
    return (datetime.now(timezone.utc) - date).days

def format_days(days):
    if days is None:
        return "Never"
    elif days == 0:
        return "Recently"
    else:
        return str(days)

def get_user_details(user):
    username = user['UserName']
    user_path = user['Path']
    details = {
        'User': username,
        'UserType': 'Service' if user_path.strip('/') == 'service' else 'Local',
        'TypeOfAccess': 'None',
        'Permissions': '',
        'LastLoginDays': None,
        'EligibleToRemove': 'No'
    }

    # Console access
    try:
        iam.get_login_profile(UserName=username)
        details['TypeOfAccess'] = 'Console'
    except iam.exceptions.NoSuchEntityException:
        pass

    # CLI access
    keys = iam.list_access_keys(UserName=username)['AccessKeyMetadata']
    if keys:
        details['TypeOfAccess'] = 'Both' if details['TypeOfAccess'] == 'Console' else 'CLI'

    # Permissions (managed + inline + group)
    policies = []

    # Direct attached managed policies
    for p in iam.list_attached_user_policies(UserName=username)['AttachedPolicies']:
        policies.append(f"Managed: {p['PolicyName']}")

    # Inline user policies
    for p in iam.list_user_policies(UserName=username)['PolicyNames']:
        policies.append(f"Inline: {p}")

    # Group-based policies
    for g in iam.list_groups_for_user(UserName=username)['Groups']:
        policies.append(f"Group: {g['GroupName']}")
        attached_group_policies = iam.list_attached_group_policies(GroupName=g['GroupName'])['AttachedPolicies']
        for gp in attached_group_policies:
            policies.append(f"GroupPolicy: {gp['PolicyName']}")

    details['Permissions'] = "\n".join([f"{i+1}. {p}" for i, p in enumerate(policies)]) if policies else "None"

    # Last login / key use
    login_date = user.get('PasswordLastUsed')
    key_used_dates = []
    for k in keys:
        info = iam.get_access_key_last_used(AccessKeyId=k['AccessKeyId'])
        last_used = info['AccessKeyLastUsed'].get('LastUsedDate')
        if last_used:
            key_used_dates.append(last_used)

    all_dates = [d for d in [login_date, *key_used_dates] if d]
    if all_dates:
        details['LastLoginDays'] = days_since(max(all_dates))

    # Removal eligibility
    if ((not login_date or days_since(login_date) >= 90)
        and (not key_used_dates or all(days_since(d) >= 90 for d in key_used_dates))):
        details['EligibleToRemove'] = 'Yes'

    return details

def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 5, 60)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(file_path)

def main():
    print("Fetching IAM users... Please wait.\n")
    paginator = iam.get_paginator('list_users')
    all_users = [get_user_details(u) for p in paginator.paginate() for u in p['Users']]

    for u in all_users:
        u['LastLoginDays'] = format_days(u['LastLoginDays'])

    headers = ["User", "UserType", "TypeOfAccess", "Permissions", "LastLoginDays", "EligibleToRemove"]
    rows = [[u[h] for h in headers] for u in all_users]
    print(tabulate(rows, headers=headers, tablefmt="grid"))
    print(f"\nTotal users: {len(all_users)}")

    output_path = os.path.expanduser("~/iam_users_report.xlsx")
    pd.DataFrame(all_users).to_excel(output_path, index=False)
    format_excel(output_path)

    print(f"\n✅ Excel report saved at: {output_path}")

if __name__ == "__main__":
    main()
